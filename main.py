# -*- coding: utf-8 -*-
import os

import kivy.resources
import pandas as pd
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.popup import Popup
from kivy.uix.filechooser import FileChooserListView
from kivy.uix.gridlayout import GridLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.textinput import TextInput
from kivy.uix.behaviors import DragBehavior
from kivy.core.window import Window
from kivy.uix.image import Image
from kivy.clock import Clock
import pytesseract
from kivy.uix.progressbar import ProgressBar
from kivy.lang import Builder
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment

Builder.load_string("""
<MainScreen>:
    orientation: 'vertical'
    BoxLayout:
        size_hint_y: 0.8
        FileChooserListView:
            id: file_chooser
            on_selection: root.load_files(self.selection)
            on_touch_down: root.on_file_chooser_touch_down(*args)
            path: './origin_data'
            font_name: 'fonts/msyh.ttc'
        ScrollView:
            GridLayout:
                id: file_list
                cols: 1
                size_hint_y: None
                height: self.minimum_height
                spacing: '5dp'
    BoxLayout:
        size_hint_y: 0.2
        Button:
            text: 'Dispose'
            on_release: root.process_files()
        Button:
            text: 'Clean'
            on_release: root.clear_data()
"""
                    )


def get_files(directory):
    file_list = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            # 构建完整的文件路径
            file_path = os.path.join(root, file)
            file_list.append(file_path)
    return file_list


class MainScreen(BoxLayout):
    def __init__(self, **kwargs):
        super(MainScreen, self).__init__(**kwargs)
        self.file_chooser = self.ids.file_chooser
        self.file_list = self.ids.file_list
        self.progress_bar = ProgressBar(max=100)
        self.progress_bar.value = 0

        self.progress_bar.size_hint_y = None
        self.progress_bar.height = '30dp'
        self.file_list.add_widget(self.progress_bar)
        self.load_files(self.file_chooser.selection)
        Window.bind(on_dropfile=self._on_file_drop)

    def load_files(self, files):
        self.file_list.clear_widgets()
        self.file_list.add_widget(self.progress_bar)
        for file in files:
            image = Image(source=file, size_hint=(1, None), height='400dp')
            self.file_list.add_widget(image)

    def process_files(self):
        # 处理origin_data目录下所有文件
        files = self.file_chooser.selection
        processed_files = 0
        if len(files) == 0:
            current_dir = os.path.dirname(os.path.abspath(__file__))
            origin_data_dir = os.path.join(current_dir, 'origin_data')
            files = get_files(origin_data_dir)
        total_files = len(files)
        print(files)
        for file in files:
            text = pytesseract.image_to_string(file, 'chi_sim')
            self.update_excel_image(f'{os.path.basename(file)}', text, file)
            processed_files += 1
            self.progress_bar.value = int(processed_files / total_files * 100)

        # 显示处理完成信息
        self.show_popup("Processing is complete")

    def update_excel_image(self, file_name, data, image_path):
        filename = 'output.xlsx'
        if os.path.isfile(filename):
            workbook = load_workbook(filename)
            sheet = workbook['Sheet1']
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = 'Sheet1'
            # 设置列宽
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 50
        sheet.column_dimensions['C'].width = 50  # 为图片留出足够的空间

        # 获取最后一行的行号
        max_row = sheet.max_row + 1

        # 插入文件路径和数据，并设置第二列数据自动换行
        sheet.cell(row=max_row, column=1, value=file_name)
        data_cell = sheet.cell(row=max_row, column=2, value=data)
        data_cell.alignment = Alignment(wrapText=True)

        # 插入图片
        if os.path.exists(image_path):
            img = OpenpyxlImage(image_path)
            # 调整图片大小以填满单元格
            # 假设图片的宽度和高度比例与单元格的比例相匹配
            img.width = sheet.column_dimensions['C'].width * 7.5  # 宽度以字符宽度为单位
            img.height = img.width  # 保持图片比例

            # 计算图片插入的位置，确保图片不重叠
            # 插入到第三列
            img_anchor = 'C{}'.format(max_row)
            sheet.add_image(img, img_anchor)

        # 保存工作簿
        workbook.save(filename)

    def update_excel(self, file_name, data):
        filename = 'output.xlsx'
        if os.path.isfile(filename):
            df = pd.DataFrame({'原始文件名': file_name, '数据': data}, index=[0])
            # 生成excel
            with pd.ExcelWriter('output.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                df.to_excel(writer, index=False, header=False, sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row)
        else:
            # 创建一个新的excel
            df = pd.DataFrame(columns=['原始文件名', '数据'])
            df = pd.concat([df, pd.DataFrame({'原始文件名': file_name, '数据': data}, index=[0])], ignore_index=True)
            with pd.ExcelWriter('output.xlsx', engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')

    def clear_data(self):
        # 删除origin_data目录
        origin_data_dir = 'origin_data'
        if os.path.exists(origin_data_dir) and os.path.isdir(origin_data_dir):
            # 遍历目录中的所有文件和子目录
            for filename in os.listdir(origin_data_dir):
                file_path = os.path.join(origin_data_dir, filename)
                try:
                    # 如果是文件，则删除
                    if os.path.isfile(file_path) or os.path.islink(file_path):
                        os.unlink(file_path)
                    # 如果是目录，则递归删除
                    elif os.path.isdir(file_path):
                        shutil.rmtree(file_path)
                except Exception as e:
                    print(f'Failed to delete {file_path}. Reason: {e}')
        if os.path.exists('output.xlsx'):
            os.remove('output.xlsx')
        # 刷新文件选择器的显示
        self.refresh_file_chooser()
        self.show_popup('Cleanup is complete')
        self.load_files([])

    def show_popup(self, message):
        layout = BoxLayout(orientation='vertical', padding='10dp')
        popup_label = Label(text=message)
        layout.add_widget(popup_label)
        popup = Popup(title='message', content=layout, size_hint=(None, None), size=('200dp', '100dp'))
        popup.open()

    def _on_file_drop(self, window, file_path):
        # 获取拖放的文件路径
        file_path = file_path.decode('utf-8')
        # 获取当前程序所在目录
        current_dir = os.path.dirname(os.path.abspath(__file__))
        origin_data_dir = os.path.join(current_dir, 'origin_data')

        # 确保origin_data目录存在
        os.makedirs(origin_data_dir, exist_ok=True)
        # 将文件复制到origin_data目录下
        shutil.copy(file_path, origin_data_dir)

        # 刷新文件选择器的显示
        self.refresh_file_chooser()

    def refresh_file_chooser(self, funcindex=0):
        # 获取当前程序所在目录
        current_dir = os.path.dirname(os.path.abspath(__file__))
        origin_data_dir = os.path.join(current_dir, 'origin_data')
        # 刷新文件选择器的显示
        if funcindex == 0:
            self.file_chooser.path = ''
        else:
            self.file_chooser.path = origin_data_dir
        Clock.schedule_once(lambda dt: setattr(self.file_chooser, 'path', origin_data_dir), 0)  # 然后在下一个帧中设置回原来的路径


    def on_file_chooser_touch_down(self, instance, touch):
        # 检查点击位置是否在文件选择器的空白处
        if instance.collide_point(*touch.pos):
            # 清除所有选中的文件
            self.file_chooser.selection = []
            # 刷新文件选择器的显示
            self.refresh_file_chooser(1)
class OCRApp(App):
    def build(self):
        Window.size = (600, 600)
        return MainScreen()


if __name__ == '__main__':
    pytesseract.pytesseract.tesseract_cmd = r'G:\jupyter\ocr_project\tesseract\tesseract.exe'
    OCRApp().run()
